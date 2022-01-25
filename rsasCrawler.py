#!/usr/bin/env python3
# -*- encoding: utf-8 -*-
"""
@Author: awei84    
@Contact: 592879940@qq.com   
@Modify_Time: 2022/1/24 10:30       
@Desciption: None
"""
# 爬取绿盟导出的报告的主机漏洞信息，代码写的比较粗放，因为直接用List存储大量数据，数据越大，越到后期，执行效率越低
# 注意：需要绿盟的报告自定义模板，只导出漏洞信息的HTML
# 导出json文件跟Excel文件

import re
import json
import openpyxl
from bs4 import BeautifulSoup
from openpyxl.styles import Font, Alignment


def crawer_vuln(file_path, vuln_num_min, vuln_num_max):
    result = []
    res_ = {
        "序号": "",
        "漏洞名称": "",
        "详细描述": "",
        "解决办法": "",
        "威胁分值": "",  # 0<x<4 低位 4<=x<7 中危 7<=x<=10高危
        "危险插件": "",
        "发现日期": "",
        "CVE编号": "",
        "CNNVD编号": "",
        "CNCVE编号": "",
        "CVSS评分": "",
        "CNVD编号": "",
        "危险等级": "",
    }
    soup = BeautifulSoup(
        open(file_path, encoding="utf-8"),
        features="html.parser",
    )  # features值可为lxml
    num_ = int(vuln_num_max) + 1
    for num in range(int(vuln_num_min), num_):
        res = res_.copy()
        table = soup.find(
            "table", attrs={"id": "vuln_distribution", "class": "report_table"}
        ).find("tbody")
        t_name = table.find(
            "tr", attrs={"onclick": f"no_toggle('1_1_{num}','table_1_1_{num}')"}
        )
        t_body = table.find("tr", attrs={"id": f"table_1_1_{num}"})
        serial_number = t_name.find_all("td")[0].text.strip()  # 序号
        vuln_name = t_name.find("span").text.strip()  # 漏洞名称
        res["序号"] = serial_number
        res["漏洞名称"] = vuln_name
        print(f"{serial_number}:{vuln_name}")

        # 详细描述
        Detail_desc = t_body.find_all("tr")[1].find("td").text.replace("\n", "").strip()
        res["详细描述"] = Detail_desc
        # print(f"详细描述：{Detail_desc}")
        Threat_score = t_body.find_all("tr")[3].find("td").text.strip()
        if Threat_score == "否":
            # 威胁分值
            Threat_score = t_body.find_all("tr")[2].find("td").text.strip()
            res["威胁分值"] = Threat_score
            print(f"威胁分值：{Threat_score}")
            if float(Threat_score) < 4:
                res["危险等级"] = "低危"
            elif 4 <= float(Threat_score) < 7:
                res["危险等级"] = "中危"
            if float(Threat_score) >= 7:
                res["危险等级"] = "高危"

            # 危险插件
            res["危险插件"] = t_body.find_all("tr")[3].find("td").text.strip()
            # print(f'危险插件：{res["危险插件"]}')

            # 发现日期
            res["发现日期"] = t_body.find_all("tr")[4].find("td").text.strip()
            # print(f'发现日期：{res["发现日期"]}')
        else:
            # 解决办法
            Solution = (
                t_body.find_all("tr")[2]
                .find("td")
                .text.replace("\n\n", "\n")
                .replace("NSFOCUS", "")
                .strip()
            )  # 没有去掉换行符
            res["解决办法"] = Solution
            # print(f"解决办法：{Solution}")

            # 威胁分值
            # Threat_score = t_body.find_all("tr")[3].find("td").text.strip()
            res["威胁分值"] = Threat_score
            print(f"威胁分值：{Threat_score}")

            if float(Threat_score) < 4:
                res["危险等级"] = "低危"
            elif 4 <= float(Threat_score) < 7:
                res["危险等级"] = "中危"
            if float(Threat_score) >= 7:
                res["危险等级"] = "高危"

            # 危险插件
            res["危险插件"] = t_body.find_all("tr")[4].find("td").text.strip()
            # print(f'危险插件：{res["危险插件"]}')

            # 发现日期
            res["发现日期"] = t_body.find_all("tr")[5].find("td").text.strip()
            # print(f'发现日期：{res["发现日期"]}')

        if float(Threat_score):  # 判断是否有错乱
            pass

        # 后面的元素不一定存在
        # CVE编号
        cve_ = t_body.find("td", text=re.compile(r"CVE-"))
        if cve_:  # 通过正则模糊查找
            res["CVE编号"] = cve_.text
        # print(f'CVE编号：{res["CVE编号"]}')

        # CNNVD编号
        CNNVD_ = t_body.find("td", text=re.compile(r"CNNVD-"))
        if CNNVD_:  # 通过正则模糊查找
            res["CNNVD编号"] = CNNVD_.text
        # print(f'CNNVD编号：{res["CNNVD编号"]}')

        # CNCVE编号
        CNCVE_ = t_body.find("td", text=re.compile(r"CNCVE-"))
        if CNCVE_:  # 通过正则模糊查找
            res["CNCVE编号"] = CNCVE_.text
        # print(f'CNCVE编号：{res["CNCVE编号"]}')

        # CNVD编号
        CNVD_ = t_body.find("td", text=re.compile(r"CNVD-"))
        if CNVD_:  # 通过正则模糊查找
            res["CNVD编号"] = CNVD_.text
        # print(f'CNVD编号：{res["CNVD编号"]}')

        # CVSS评分
        CVSS_ = t_body.find_all("tr")
        for x in CVSS_:
            if "CVSS评分" in x.text:
                res["CVSS评分"] = x.find("td").text
                break
        # print(f'CVSS评分：{res["CVSS评分"]}\n')
        result.append(res)
    return result


def save_excel(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws1 = wb.create_sheet("VulnPlugDesc", 0)
    titlelist = list(data[0].keys())
    ws1.append(titlelist)
    for x in titlelist:  # title
        ws1.cell(column=(int(titlelist.index(x)) + 1), row=1).value = x
        ws1.cell(column=(int(titlelist.index(x)) + 1), row=1).font = Font(bold=True)
        ws1.cell(column=(int(titlelist.index(x)) + 1), row=1).alignment = Alignment(
            horizontal="center", vertical="center"
        )
    for v in data:
        ws1.append(list(v.values()))

    wb.save(r"C:\Users\59287\Desktop\绿盟插件表\绿盟漏洞插件表.xlsx")


def save_json(data):
    data_j = json.dumps(data,indent=4)
    with open(r"C:\Users\59287\Desktop\绿盟插件表\vuln.json","w") as f:
        f.write(data_j)

file_path = "D:\Download\多任务输出_2022_01_25_html\index.html"
data = crawer_vuln(file_path, "1", "2298") # 第1条到2298条
save_json(data)
save_excel(data)
